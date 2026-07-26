#!/usr/bin/env python3
"""
ats_performance_report.py

Generates a TradeStation-style performance report from the AtsFastReversal
(or AtsSlowReversal) merged CSV trade file.

Usage:
    python ats_performance_report.py <csv_file> [--strategy NAME] [--excel FILE] [--output FILE]

Output:
    Console report + optional Excel file with three sheets:
      - Trade Summary: Key metrics and statistics
      - Daily Performance: Daily P/L and win rates
      - Performance By Symbol: Per-symbol statistics
    OR optional CSV summary (legacy)

All raw CSV column names live in the shared `Column` enum
(optimizer_constants.py). All the core logic here is consolidated into the
`PerformanceReportGenerator` class.
"""

import sys
import argparse
from enum import Enum

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from optimizer_constants import Column, PROFIT_HIT_COL


# ============================================================================
# WorkingColumn -- derived/working column names computed by this script
# (as opposed to Column, which is raw CSV source columns). Kept local to
# this file since these are report-internal, not part of the shared CSV
# schema.
# ============================================================================
class WorkingColumn(str, Enum):
    ENTRY_DT = "EntryDT"
    EXIT_DT = "ExitDT"
    PL = "PL"
    IS_LONG = "IsLong"
    WIN = "Win"
    DATE = "Date"
    POSITION_VALUE = "PositionValue"


# ============================================================================
# PerformanceReportGenerator -- consolidates all the logic used to load
# trades, compute performance metrics, print the console report, and build
# the multi-sheet Excel report.
# ============================================================================
class PerformanceReportGenerator:

    def __init__(self, strategy_name: str = None):
        self.strategy_name = strategy_name

    # ------------------------------------------------------------------
    # Data loading / filtering
    # ------------------------------------------------------------------
    @staticmethod
    def load_trades(csv_path: str) -> pd.DataFrame:
        df = pd.read_csv(csv_path)
        df.columns = [c.strip() for c in df.columns]

        # Parse dates / times
        df[WorkingColumn.ENTRY_DT.value] = pd.to_datetime(
            df[Column.ENTRY_DATE.value].astype(str) + " " +
            df[Column.ENTRY_TIME.value].astype(str).str.zfill(4).str[:2] + ":" +
            df[Column.ENTRY_TIME.value].astype(str).str.zfill(4).str[2:],
            format="%m/%d/%Y %H:%M", errors="coerce"
        )
        df[WorkingColumn.EXIT_DT.value] = pd.to_datetime(
            df[Column.EXIT_DATE.value].astype(str) + " " +
            df[Column.EXIT_TIME.value].astype(str).str.zfill(4).str[:2] + ":" +
            df[Column.EXIT_TIME.value].astype(str).str.zfill(4).str[2:],
            format="%m/%d/%Y %H:%M", errors="coerce"
        )

        df[WorkingColumn.PL.value] = pd.to_numeric(df[Column.PROFIT_LOSS.value], errors="coerce").fillna(0)
        df[Column.SHARES.value] = pd.to_numeric(df[Column.SHARES.value], errors="coerce").fillna(0)
        df[WorkingColumn.IS_LONG.value] = (
            df[Column.IND_SIGNAL_SENT.value].eq(1) if Column.IND_SIGNAL_SENT.value in df.columns
            else df[Column.ENTRY_NAME.value].str.startswith("LE")
        )
        df[WorkingColumn.WIN.value] = df[WorkingColumn.PL.value] > 0
        df[PROFIT_HIT_COL] = df[WorkingColumn.PL.value] > 0  # at least one target hit (see context doc)

        return df

    @staticmethod
    def filter_trades(df: pd.DataFrame, direction: str = "both", symbol: str = None,
                       start: str = None, end: str = None) -> pd.DataFrame:
        if direction == "long":
            df = df[df[WorkingColumn.IS_LONG.value]]
        elif direction == "short":
            df = df[~df[WorkingColumn.IS_LONG.value]]
        if symbol:
            df = df[df[Column.SYMBOL.value].str.upper() == symbol.upper()]
        if start:
            df = df[df[WorkingColumn.ENTRY_DT.value] >= pd.to_datetime(start, format="%m/%d/%Y")]
        if end:
            df = df[df[WorkingColumn.ENTRY_DT.value] <= pd.to_datetime(end, format="%m/%d/%Y")]
        return df.copy()

    # ------------------------------------------------------------------
    # Core metrics
    # ------------------------------------------------------------------
    @staticmethod
    def calc_equity_curve(pl_series: pd.Series) -> pd.Series:
        """Cumulative P/L series."""
        return pl_series.cumsum()

    @staticmethod
    def calc_drawdown(equity: pd.Series) -> tuple:
        """Returns drawdown series, max drawdown value, and max drawdown duration (bars)."""
        roll_max = equity.cummax()
        dd = equity - roll_max                 # always <= 0
        max_dd = dd.min()

        # Duration: longest consecutive period underwater
        underwater = (dd < 0).astype(int)
        max_dur, cur_dur = 0, 0
        for v in underwater:
            if v:
                cur_dur += 1
                max_dur = max(max_dur, cur_dur)
            else:
                cur_dur = 0

        return dd, max_dd, max_dur

    @staticmethod
    def calc_sharpe(pl_series: pd.Series, periods_per_year: int = 252) -> float:
        """Daily Sharpe ratio (assumes each row = 1 trading day unit; adjust as needed)."""
        if pl_series.std() == 0:
            return np.nan
        return (pl_series.mean() / pl_series.std()) * np.sqrt(periods_per_year)

    @staticmethod
    def calc_sortino(pl_series: pd.Series, periods_per_year: int = 252) -> float:
        """Sortino ratio using downside deviation."""
        neg = pl_series[pl_series < 0]
        if len(neg) == 0 or neg.std() == 0:
            return np.nan
        return (pl_series.mean() / neg.std()) * np.sqrt(periods_per_year)

    @staticmethod
    def calc_profit_factor(pl_series: pd.Series) -> float:
        gross_win = pl_series[pl_series > 0].sum()
        gross_loss = abs(pl_series[pl_series < 0].sum())
        return gross_win / gross_loss if gross_loss > 0 else np.inf

    @staticmethod
    def calc_expectancy(pl_series: pd.Series) -> float:
        wins = pl_series[pl_series > 0]
        losses = pl_series[pl_series < 0]
        wr = len(wins) / len(pl_series) if len(pl_series) > 0 else 0
        avg_w = wins.mean() if len(wins) > 0 else 0
        avg_l = losses.mean() if len(losses) > 0 else 0
        return wr * avg_w + (1 - wr) * avg_l

    @staticmethod
    def calc_max_capital(df: pd.DataFrame) -> float:
        """
        Approximate max capital required:
        max(EntryPrice * Shares) across all trades -- the largest single
        position cost. For a more accurate margin calc you'd need the
        broker margin rate.
        """
        if Column.ENTRY_PRICE.value in df.columns:
            df2 = df.copy()
            df2[WorkingColumn.POSITION_VALUE.value] = (
                pd.to_numeric(df2[Column.ENTRY_PRICE.value], errors="coerce") * df2[Column.SHARES.value]
            )
            return df2[WorkingColumn.POSITION_VALUE.value].max()
        return np.nan

    @staticmethod
    def calc_consecutive(win_series) -> tuple:
        """Max consecutive wins and losses."""
        max_w = max_l = cur_w = cur_l = 0
        for w in win_series:
            if w:
                cur_w += 1
                cur_l = 0
            else:
                cur_l += 1
                cur_w = 0
            max_w = max(max_w, cur_w)
            max_l = max(max_l, cur_l)
        return max_w, max_l

    @staticmethod
    def exit_breakdown(df: pd.DataFrame) -> pd.DataFrame:
        return df.groupby(Column.EXIT_NAME.value).agg(
            Count=(WorkingColumn.PL.value, "size"),
            TotalPL=(WorkingColumn.PL.value, "sum"),
            AvgPL=(WorkingColumn.PL.value, "mean"),
            WinRate=(WorkingColumn.WIN.value, "mean"),
        ).sort_values("Count", ascending=False)

    @staticmethod
    def daily_stats(df: pd.DataFrame) -> pd.DataFrame:
        df2 = df.copy()
        df2[WorkingColumn.DATE.value] = df2[WorkingColumn.ENTRY_DT.value].dt.date
        return df2.groupby(WorkingColumn.DATE.value).agg(
            Trades=(WorkingColumn.PL.value, "size"),
            TotalPL=(WorkingColumn.PL.value, "sum"),
            WinRate=(WorkingColumn.WIN.value, "mean"),
        )

    @staticmethod
    def symbol_stats(df: pd.DataFrame) -> pd.DataFrame:
        return df.groupby(Column.SYMBOL.value).agg(
            Trades=(WorkingColumn.PL.value, "size"),
            TotalPL=(WorkingColumn.PL.value, "sum"),
            AvgPL=(WorkingColumn.PL.value, "mean"),
            WinRate=(WorkingColumn.WIN.value, "mean"),
        ).sort_values("TotalPL", ascending=False)

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------
    @staticmethod
    def hr(char: str = "\u2500", width: int = 64):
        print(char * width)

    @staticmethod
    def fmt(val, prefix: str = "$", decimals: int = 2) -> str:
        if pd.isna(val):
            return "N/A"
        if val == np.inf:
            return "\u221e"
        return f"{prefix}{val:,.{decimals}f}" if prefix else f"{val:.{decimals}f}"

    @staticmethod
    def pct(val) -> str:
        return "N/A" if pd.isna(val) else f"{val*100:.1f}%"

    # ------------------------------------------------------------------
    # Console report
    # ------------------------------------------------------------------
    def print_report(self, df: pd.DataFrame, strategy_name: str) -> dict:
        pl = df[WorkingColumn.PL.value]
        wins = df[df[WorkingColumn.WIN.value]]
        losses = df[~df[WorkingColumn.WIN.value]]
        longs = df[df[WorkingColumn.IS_LONG.value]]
        shorts = df[~df[WorkingColumn.IS_LONG.value]]

        equity = self.calc_equity_curve(pl.reset_index(drop=True))
        dd, max_dd, max_dd_dur = self.calc_drawdown(equity)
        max_cap = self.calc_max_capital(df)
        max_w, max_l = self.calc_consecutive(df[WorkingColumn.WIN.value].tolist())

        # Daily P/L for Sharpe (group by date)
        df2 = df.copy()
        df2[WorkingColumn.DATE.value] = df2[WorkingColumn.ENTRY_DT.value].dt.date
        daily_pl = df2.groupby(WorkingColumn.DATE.value)[WorkingColumn.PL.value].sum()

        fmt, pct, hr = self.fmt, self.pct, self.hr

        print()
        hr("\u2550")
        print(f"  PERFORMANCE REPORT \u2014 {strategy_name}")
        hr("\u2550")

        # -- Summary --
        print("\n  TRADE SUMMARY")
        hr()
        print(f"  {'Total Net Profit':<35} {fmt(pl.sum())}")
        print(f"  {'Gross Profit':<35} {fmt(pl[pl>0].sum())}")
        print(f"  {'Gross Loss':<35} {fmt(pl[pl<0].sum())}")
        print(f"  {'Profit Factor':<35} {fmt(self.calc_profit_factor(pl), prefix='', decimals=3)}")
        print(f"  {'Expectancy (per trade)':<35} {fmt(self.calc_expectancy(pl))}")
        print()
        print(f"  {'Total Trades':<35} {len(df)}")
        print(f"  {'  Winning Trades':<35} {len(wins)}  ({pct(len(wins)/len(df) if df.size else 0)})")
        print(f"  {'  Losing Trades':<35} {len(losses)}  ({pct(len(losses)/len(df) if df.size else 0)})")
        print(f"  {'  Scratch (P/L = 0)':<35} {len(df[pl==0])}")
        print()
        print(f"  {'Long Trades':<35} {len(longs)}  ({pct(len(longs)/len(df) if df.size else 0)})")
        print(f"  {'  Long Win Rate':<35} {pct(longs[WorkingColumn.WIN.value].mean())}")
        print(f"  {'  Long Net P/L':<35} {fmt(longs[WorkingColumn.PL.value].sum())}")
        print(f"  {'Short Trades':<35} {len(shorts)}  ({pct(len(shorts)/len(df) if df.size else 0)})")
        print(f"  {'  Short Win Rate':<35} {pct(shorts[WorkingColumn.WIN.value].mean())}")
        print(f"  {'  Short Net P/L':<35} {fmt(shorts[WorkingColumn.PL.value].sum())}")

        # -- Win/Loss Detail --
        print("\n  WIN / LOSS DETAIL")
        hr()
        print(f"  {'Avg Winning Trade':<35} {fmt(wins[WorkingColumn.PL.value].mean())}")
        print(f"  {'Avg Losing Trade':<35} {fmt(losses[WorkingColumn.PL.value].mean())}")
        print(f"  {'Largest Win':<35} {fmt(wins[WorkingColumn.PL.value].max())}")
        print(f"  {'Largest Loss':<35} {fmt(losses[WorkingColumn.PL.value].min())}")
        avg_w_l_ratio = (abs(wins[WorkingColumn.PL.value].mean() / losses[WorkingColumn.PL.value].mean())
                         if losses[WorkingColumn.PL.value].mean() != 0 else np.nan)
        print(f"  {'Avg Win / Avg Loss Ratio':<35} {fmt(avg_w_l_ratio, prefix='', decimals=3)}")
        print(f"  {'Max Consecutive Winners':<35} {max_w}")
        print(f"  {'Max Consecutive Losers':<35} {max_l}")
        print(f"  {'Avg Shares per Trade':<35} {fmt(df[Column.SHARES.value].mean(), prefix='', decimals=1)}")

        # -- Drawdown & Capital --
        print("\n  DRAWDOWN & CAPITAL")
        hr()
        print(f"  {'Max Drawdown (P/L units)':<35} {fmt(max_dd)}")
        print(f"  {'Max Drawdown Duration (trades)':<35} {max_dd_dur}")
        print(f"  {'Max Capital Required (est.)':<35} {fmt(max_cap)}")
        print(f"  {'Return on Max Capital':<35} {pct(pl.sum()/max_cap if max_cap and max_cap>0 else np.nan)}")

        # -- Risk Ratios --
        print("\n  RISK-ADJUSTED METRICS  (annualised, 252 trading days)")
        hr()
        print(f"  {'Sharpe Ratio':<35} {fmt(self.calc_sharpe(daily_pl), prefix='', decimals=3)}")
        print(f"  {'Sortino Ratio':<35} {fmt(self.calc_sortino(daily_pl), prefix='', decimals=3)}")
        print(f"  {'Calmar Ratio (Net/MaxDD)':<35} {fmt(pl.sum()/abs(max_dd) if max_dd!=0 else np.nan, prefix='', decimals=3)}")

        # -- Exit Breakdown --
        print("\n  EXIT TYPE BREAKDOWN")
        hr()
        eb = self.exit_breakdown(df)
        print(f"  {'Exit':<20} {'Count':>6} {'TotalPL':>10} {'AvgPL':>8} {'WinRate':>8}")
        hr()
        for name, erow in eb.iterrows():
            print(f"  {name:<20} {int(erow.Count):>6} {fmt(erow.TotalPL):>10} {fmt(erow.AvgPL):>8} {pct(erow.WinRate):>8}")

        # -- Daily --
        print("\n  DAILY PERFORMANCE")
        hr()
        ds = self.daily_stats(df)
        print(f"  {'Date':<14} {'Trades':>6} {'P/L':>10} {'WinRate':>8}")
        hr()
        for date, drow in ds.iterrows():
            print(f"  {str(date):<14} {int(drow.Trades):>6} {fmt(drow.TotalPL):>10} {pct(drow.WinRate):>8}")
        print(f"\n  {'Best Day':<35} {fmt(ds['TotalPL'].max())}")
        print(f"  {'Worst Day':<35} {fmt(ds['TotalPL'].min())}")
        print(f"  {'Avg Daily P/L':<35} {fmt(ds['TotalPL'].mean())}")
        print(f"  {'Profitable Days':<35} {(ds['TotalPL']>0).sum()} / {len(ds)}")

        # -- By Symbol --
        if df[Column.SYMBOL.value].nunique() > 1:
            print("\n  PERFORMANCE BY SYMBOL")
            hr()
            ss = self.symbol_stats(df)
            print(f"  {'Symbol':<10} {'Trades':>6} {'TotalPL':>10} {'AvgPL':>8} {'WinRate':>8}")
            hr()
            for sym, srow in ss.iterrows():
                print(f"  {sym:<10} {int(srow.Trades):>6} {fmt(srow.TotalPL):>10} {fmt(srow.AvgPL):>8} {pct(srow.WinRate):>8}")

        hr("\u2550")
        print()

        return {
            "TotalTrades": len(df),
            "WinRate": df[WorkingColumn.WIN.value].mean(),
            "NetPL": pl.sum(),
            "GrossProfit": pl[pl>0].sum(),
            "GrossLoss": pl[pl<0].sum(),
            "ProfitFactor": self.calc_profit_factor(pl),
            "Expectancy": self.calc_expectancy(pl),
            "AvgWin": wins[WorkingColumn.PL.value].mean(),
            "AvgLoss": losses[WorkingColumn.PL.value].mean(),
            "LargestWin": wins[WorkingColumn.PL.value].max(),
            "LargestLoss": losses[WorkingColumn.PL.value].min(),
            "MaxDrawdown": max_dd,
            "MaxDrawdownDuration": max_dd_dur,
            "MaxCapitalRequired": max_cap,
            "Sharpe": self.calc_sharpe(daily_pl),
            "Sortino": self.calc_sortino(daily_pl),
            "MaxConsecWins": max_w,
            "MaxConsecLosses": max_l,
        }

    # ------------------------------------------------------------------
    # Excel report
    # ------------------------------------------------------------------
    def create_excel_report(self, df: pd.DataFrame, strategy_name: str, output_path: str):
        """Create an Excel workbook with three sheets: Trade Summary, Daily Performance, Performance By Symbol."""
        fmt, pct = self.fmt, self.pct

        pl = df[WorkingColumn.PL.value]
        wins = df[df[WorkingColumn.WIN.value]]
        losses = df[~df[WorkingColumn.WIN.value]]
        longs = df[df[WorkingColumn.IS_LONG.value]]
        shorts = df[~df[WorkingColumn.IS_LONG.value]]

        equity = self.calc_equity_curve(pl.reset_index(drop=True))
        dd, max_dd, max_dd_dur = self.calc_drawdown(equity)
        max_cap = self.calc_max_capital(df)
        max_w, max_l = self.calc_consecutive(df[WorkingColumn.WIN.value].tolist())

        df2 = df.copy()
        df2[WorkingColumn.DATE.value] = df2[WorkingColumn.ENTRY_DT.value].dt.date
        daily_pl = df2.groupby(WorkingColumn.DATE.value)[WorkingColumn.PL.value].sum()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Summary"

        # -- Sheet 1: Trade Summary --
        row = 1

        def add_section(title, start_row):
            ws[f"A{start_row}"] = title
            ws[f"A{start_row}"].font = Font(bold=True, size=12)
            ws[f"A{start_row}"].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            return start_row + 1

        row = add_section("TRADE SUMMARY", row)
        summary_data = [
            ("Total Net Profit", fmt(pl.sum())),
            ("Gross Profit", fmt(pl[pl>0].sum())),
            ("Gross Loss", fmt(pl[pl<0].sum())),
            ("Profit Factor", fmt(self.calc_profit_factor(pl), prefix='', decimals=3)),
            ("Expectancy (per trade)", fmt(self.calc_expectancy(pl))),
        ]
        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        row += 1

        row = add_section("TRADE COUNT", row)
        trade_data = [
            ("Total Trades", len(df)),
            ("Winning Trades", f"{len(wins)} ({pct(len(wins)/len(df) if df.size else 0)})"),
            ("Losing Trades", f"{len(losses)} ({pct(len(losses)/len(df) if df.size else 0)})"),
            ("Scratch (P/L = 0)", len(df[pl==0])),
            ("Long Trades", f"{len(longs)} ({pct(len(longs)/len(df) if df.size else 0)})"),
            ("Long Win Rate", pct(longs[WorkingColumn.WIN.value].mean())),
            ("Long Net P/L", fmt(longs[WorkingColumn.PL.value].sum())),
            ("Short Trades", f"{len(shorts)} ({pct(len(shorts)/len(df) if df.size else 0)})"),
            ("Short Win Rate", pct(shorts[WorkingColumn.WIN.value].mean())),
            ("Short Net P/L", fmt(shorts[WorkingColumn.PL.value].sum())),
        ]
        for label, value in trade_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        row += 1

        row = add_section("WIN / LOSS DETAIL", row)
        avg_w_l_ratio = (abs(wins[WorkingColumn.PL.value].mean() / losses[WorkingColumn.PL.value].mean())
                         if losses[WorkingColumn.PL.value].mean() != 0 else np.nan)
        winloss_data = [
            ("Avg Winning Trade", fmt(wins[WorkingColumn.PL.value].mean())),
            ("Avg Losing Trade", fmt(losses[WorkingColumn.PL.value].mean())),
            ("Largest Win", fmt(wins[WorkingColumn.PL.value].max())),
            ("Largest Loss", fmt(losses[WorkingColumn.PL.value].min())),
            ("Avg Win / Avg Loss Ratio", fmt(avg_w_l_ratio, prefix='', decimals=3)),
            ("Max Consecutive Winners", max_w),
            ("Max Consecutive Losers", max_l),
            ("Avg Shares per Trade", fmt(df[Column.SHARES.value].mean(), prefix='', decimals=1)),
        ]
        for label, value in winloss_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        row += 1

        row = add_section("DRAWDOWN & CAPITAL", row)
        dd_data = [
            ("Max Drawdown (P/L units)", fmt(max_dd)),
            ("Max Drawdown Duration (trades)", max_dd_dur),
            ("Max Capital Required (est.)", fmt(max_cap)),
            ("Return on Max Capital", pct(pl.sum()/max_cap if max_cap and max_cap>0 else np.nan)),
        ]
        for label, value in dd_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        row += 1

        row = add_section("RISK-ADJUSTED METRICS (annualised, 252 trading days)", row)
        risk_data = [
            ("Sharpe Ratio", fmt(self.calc_sharpe(daily_pl), prefix='', decimals=3)),
            ("Sortino Ratio", fmt(self.calc_sortino(daily_pl), prefix='', decimals=3)),
            ("Calmar Ratio (Net/MaxDD)", fmt(pl.sum()/abs(max_dd) if max_dd!=0 else np.nan, prefix='', decimals=3)),
        ]
        for label, value in risk_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

        # -- Sheet 2: Daily Performance --
        ws_daily = wb.create_sheet("Daily Performance")
        ds = self.daily_stats(df)
        ds_reset = ds.reset_index()

        headers = ["Date", "Trades", "Total P/L", "Win Rate"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_daily.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for row_num, row_data in enumerate(dataframe_to_rows(ds_reset, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_daily.cell(row=row_num, column=col_num)
                cell.value = value

        ws_daily.column_dimensions['A'].width = 15
        ws_daily.column_dimensions['B'].width = 12
        ws_daily.column_dimensions['C'].width = 15
        ws_daily.column_dimensions['D'].width = 12

        # -- Sheet 3: Performance By Symbol --
        ws_symbol = wb.create_sheet("Performance By Symbol")
        if df[Column.SYMBOL.value].nunique() > 1:
            ss = self.symbol_stats(df)
            ss_reset = ss.reset_index()

            headers = ["Symbol", "Trades", "Total P/L", "Avg P/L", "Win Rate"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_symbol.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for row_num, row_data in enumerate(dataframe_to_rows(ss_reset, index=False, header=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_symbol.cell(row=row_num, column=col_num)
                    cell.value = value

            ws_symbol.column_dimensions['A'].width = 12
            ws_symbol.column_dimensions['B'].width = 12
            ws_symbol.column_dimensions['C'].width = 15
            ws_symbol.column_dimensions['D'].width = 15
            ws_symbol.column_dimensions['E'].width = 12
        else:
            ws_symbol["A1"] = "Only one symbol in dataset"

        wb.save(output_path)


def parse_args():
    p = argparse.ArgumentParser(description="ATS Performance Report")
    p.add_argument("csv_file", help="Path to merged trade CSV")
    p.add_argument("--strategy", default=None, help="Strategy name override")
    p.add_argument("--output", default=None, help="Save summary to CSV file (deprecated, use --excel)")
    p.add_argument("--excel", default=None, help="Save report to Excel file with multiple sheets")
    p.add_argument("--direction", choices=["long", "short", "both"], default="both",
                   help="Filter by trade direction")
    p.add_argument("--symbol", default=None, help="Filter by symbol")
    p.add_argument("--start", default=None, help="Start date mm/dd/yyyy")
    p.add_argument("--end", default=None, help="End date   mm/dd/yyyy")
    return p.parse_args()


def main():
    args = parse_args()
    generator = PerformanceReportGenerator(strategy_name=args.strategy)

    df = generator.load_trades(args.csv_file)
    df = generator.filter_trades(df, direction=args.direction, symbol=args.symbol,
                                  start=args.start, end=args.end)

    if df.empty:
        print("No trades match the specified filters.")
        sys.exit(1)

    strategy = args.strategy or args.csv_file.split("/")[-1].replace(".csv", "")
    summary = generator.print_report(df, strategy)

    # Handle Excel output (preferred)
    if args.excel:
        generator.create_excel_report(df, strategy, args.excel)
        print(f"Excel report saved to {args.excel}")

    # Handle legacy CSV output
    if args.output:
        pd.DataFrame([summary]).to_csv(args.output, index=False)
        print(f"Summary saved to {args.output}")


if __name__ == "__main__":
    main()
